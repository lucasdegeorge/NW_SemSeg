#%%
import json
import os
import torchvision.transforms as T
import cv2
import PIL
from PIL import Image
import openpyxl
# import sys
import matplotlib.pyplot as plt
import pandas as pd
import glob

# sys.path.append("C:/Users/lucas/Documents/GitHub/NW_SemSeg")
from CCT_models.CCT_models import *
from unet_models import *
from measurement import *
from display import *

def get_column_index(sheet, column_name):
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value == column_name:
            return col_idx
    return None


class Analysis():
    def __init__(self, models_path, images_path, videos_path, result_file=None):
        """ models_path: a path or a list of path of .pth files (representating models)
            images_path: a path of an image or a folder with .png images to analyse
            videos_path: a path of a video or a folder with .avi videos to analyse """
        
        # read jsons
        with open("parameters/CCT_parameters.json", 'r') as f:
            self.CCT_arguments = json.load(f)
            self.device = self.CCT_arguments["device"]
        with open("parameters/UNet_parameters.json", 'r') as f:
            self.UNet_arguments = json.load(f)
        with open("models_info.json", 'r') as f:
            self.model_dict = json.load(f)
        with open("analysis_parameters.json", 'r') as f:
            self.analysis_param = json.load(f)

        self.result_file = result_file
        self.preloading = False

        # images and models
        self.models, self.models_names = self.load_models(models_path)
        assert (images_path is None)^(videos_path is None), "Choose one and only one of the image and video modes"
        if images_path is not None:  self.images = self.load_images(images_path)
        if videos_path is not None:  self.images = self.load_videos(videos_path)
    
    def load_models(self, models_path): 
        model_list = []
        model_name_list = []

        for model_name in models_path:
            model_path = model_folder + "/" + model_name + "/" + model_name + "_best.pth"

            if "model" in model_name:
                CCT = True
                if "semi" in model_name:
                    model = Model(mode="semi")
                    with open(model_path, 'rb') as f:
                        model.load_state_dict(torch.load(model_path, map_location=device))
                elif "super" in model_name:
                    model = Model(mode="super")
                    with open(model_path, 'rb') as f:
                        model.load_state_dict(torch.load(model_path, map_location=device))

            elif "unet" in model_name:
                CCT = False
                if "semi" in model_name:
                    model_path = model_path.replace('semi_', '')
                    model = ssl_UNet(self.UNet_arguments)
                    with open(model_path, 'rb') as f:
                        model.load_state_dict(torch.load(model_path, map_location=device))

                elif "super" in model_name:
                    model_path = model_path.replace('super_', '')
                    model = UNet(self.UNet_arguments)
                    with open(model_path, 'rb') as f:
                        model.load_state_dict(torch.load(model_path, map_location=device))
            
            model_list.append(model)
            model_name_list.append(model_name)

        return model_list, model_name_list
    
    def load_images(self, images_path, main_dir=True):
        if images_path.endswith(".png") or images_path.endswith(".pt"):
            if main_dir:
                os.makedirs(images_path.split(".png")[0] + "_analysis", exist_ok=True)
                self.main_analysis_dir = images_path.split(".png")[0] + "_analysis"
            return [images_path]
        else:
            if main_dir:
                os.makedirs(images_path + "/analysis", exist_ok=True)
                self.main_analysis_dir = images_path + "/analysis"
            return glob.glob(images_path + '/*.png')
        
    def load_videos(self, videos_path):
        if videos_path.endswith(".avi"):
            # folder for the frames extracted
            output_folder_name = videos_path.split(".avi")[0] + "_frames"
            os.makedirs(output_folder_name, exist_ok=True)
            # folder for the analysis 
            os.makedirs(videos_path.split(".avi")[0] + "_analysis", exist_ok=True)
            self.main_analysis_dir = videos_path.split(".avi")[0] + "_analysis"
            self.extract_frames(videos_path, output_folder=output_folder_name)
            return self.load_images(output_folder_name, main_dir=False)
        else:
            # folders for the frames extracted
            main_output_folder_name = videos_path + "/frames"
            os.makedirs(main_output_folder_name, exist_ok=True)
            # folder for the analysis 
            os.makedirs(videos_path + "/analysis", exist_ok=True)
            self.main_analysis_dir = videos_path + "/analysis"
            
            files = os.listdir(videos_path)
            video_list = []
            for video in files:
                if not(video.endswith(".avi")): 
                    continue
                output_folder_name = os.path.join(main_output_folder_name, video.split(".avi")[0] + "_frames")
                self.extract_frames(os.path.join(videos_path, video), output_folder=output_folder_name)
                video_list.append(self.load_images(output_folder_name, main_dir=False))
            return [frame for sublist in video_list for frame in sublist]  # flatten video_list (which was a list of lists)

    def extract_frames(self, video_file, output_folder, output_format='png'):
        print("extracting video", video_file)
        if not(video_file.endswith('.avi')): 
            raise ValueError("The video must be an .avi video")
        
        video = cv2.VideoCapture(video_file)
        frame_count = 0
        while video.isOpened():
            # Read the next frame
            ret, frame = video.read()
            # Check if the frame was read successfully
            if ret:
                os.makedirs(output_folder, exist_ok=True)  
                output_file = os.path.join(output_folder, video_file.split("/")[-1].split("\\")[-1].split(".")[0] + '_frame_' + '{:0>5}'.format(frame_count) + '.' + output_format)
                cv2.imwrite(output_file, frame)
                # print(f'Saved frame {frame_count} as {output_file}')
                frame_count += 1
                if frame_count % 100 == 0:
                    print("current frame:", frame_count )
            else:
                break
        video.release()

    def get_name_image(self, image_idx):
        # print(self.images[image_idx])
        if "\\" in self.images[image_idx]:
            # print(self.images[image_idx].split("\\")[-1].split(".")[0])
            return self.images[image_idx].split("\\")[-1].split(".")[0]
        else:
            # print(self.images[image_idx].split("/")[-1].split(".")[0])
            return self.images[image_idx].split("/")[-1].split(".")[0]

    def save_image(self, array, save_name):
        array = Image.fromarray(array.astype(np.uint8), mode='L')
        array.save(save_name)

    def predict_1image_1model(self, image_idx, model_idx, save_display=True, return_image=False):
        model = self.models[model_idx]
        model.eval()
        model.to(self.device)

        # load the image
        # converter = T.ToTensor()
        # image = Image.open(self.images[image_idx]).convert('L')
        # image = image.resize((512,512), resample=PIL.Image.Resampling.NEAREST)
        # image = converter(image).to(self.device)
        image = torch.load(self.images[image_idx])

        # predict
        if self.model_dict[self.models_names[model_idx]][0] == "CCT": 
            prediction = model(image.unsqueeze(0).to(device), eval=True)["output_l"][0]
        else: 
            prediction = model(image.unsqueeze(0).to(device), eval=True)[0]
        prediction = prediction.permute(1,2,0)
        prediction = torch.softmax(prediction, dim=-1)
        prediction = torch.argmax(prediction, dim=-1)

        # display and save
        if save_display:
            os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]), exist_ok=True)

            if self.analysis_param["predict"]["save_pred"]:
                os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/pred", exist_ok=True)
                grayscale_values = {0: 0, 1: 127, 2: 255}
                pred_array = np.vectorize(grayscale_values.get)(prediction.cpu().numpy())
                save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/pred/" + self.get_name_image(image_idx) +  "_pred.png"
                self.save_image(pred_array, save_name)
            
            if self.analysis_param["predict"]["display_pred"]:
                grayscale_values = {0: 0, 1: 127, 2: 255}
                pred_array = np.vectorize(grayscale_values.get)(prediction.cpu().numpy())
                plt.figure()
                plt.imshow(pred_array)
                plt.plot()

            if self.analysis_param["predict"]["save_image_with_mask"] or self.analysis_param["predict"]["display_image_with_mask"]:
                os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/image_with_mask", exist_ok=True)
                save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/image_with_mask/" + self.get_name_image(image_idx) + "_im_pred.png"
                display_save_image_with_mask(image, prediction, save=self.analysis_param["predict"]["save_image_with_mask"], display=self.analysis_param["predict"]["display_image_with_mask"], save_name=save_name)

            if self.analysis_param["predict"]["save_overlaid"] or self.analysis_param["predict"]["display_overlaid"]:
                os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/overlaid", exist_ok=True)
                save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/overlaid/" + self.get_name_image(image_idx) + "overlaid.png"
                display_save_image_mask_overlaid(image, prediction, save=self.analysis_param["predict"]["save_overlaid"], display=self.analysis_param["predict"]["display_overlaid"], save_name=save_name)

        if return_image:
            return prediction, image
        else:
            return prediction
        
    def predict(self, model_idx):
        for idx in range(len(self.images)):
            self.predict_1image_1model(model_idx, idx)

    def get_only_droplet_edge(self, image_idx, model_idx):
        pred = self.predict_1image_1model(model_idx, image_idx, save_display=False).cpu().numpy()

        pred_wire = np.zeros_like(pred).astype('uint8')
        pred_droplet = np.zeros_like(pred).astype('uint8')
        pred_wire[pred == 1] = 1
        pred_droplet[pred == 2] = 1

        droplet_edge = erode_and_dilate(pred_droplet)
        wire_edge = erode_and_dilate(pred_wire)
        interface = get_interface(wire_edge, droplet_edge)

        only_droplet_edge = cv2.bitwise_xor(droplet_edge, interface, mask=droplet_edge)

        return only_droplet_edge

    def get_circle(self, image_idx, model_idx, save_display=False):
        # prediction 
        only_droplet_edge = self.get_only_droplet_edge(model_idx, image_idx)

        # find circle
        circle_info = cv2.HoughCircles(only_droplet_edge, cv2.HOUGH_GRADIENT,2,500, param1=300,param2=0.9,minRadius=0,maxRadius=0)[0,0].astype(np.int16)
        R = circle_info[2]
        rr, cc = circle_perimeter(circle_info[1], circle_info[0], circle_info[2])

        # save and display 
        if save_display:
            if self.analysis_param["get_circle"]["save_with_edges"] or self.analysis_param["get_circle"]["display_with_edges"]:

                # image creation 
                droplet_with_circle = copy.deepcopy(only_droplet_edge)
                droplet_with_circle[rr, cc] = 128

                if self.analysis_param["get_circle"]["save_with_edges"]:
                    os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/circle", exist_ok=True)
                    save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/circle/" + self.get_name_image(image_idx) +  "_circle.png"
                    self.save_image(droplet_with_circle, save_name)
                if self.analysis_param["get_circle"]["display_with_edges"]:
                    plt.figure()
                    plt.imshow(droplet_with_circle) 

        return R, circle_info[0], circle_info[1]

    def get_ellipse(self, image_idx, model_idx):
        ''' It returns a tuple of tuples in ((x,y), (majorAxis, minorAxis), angle) '''
        # prediction: 
        only_droplet_edge = self.get_only_droplet_edge(model_idx, image_idx)

        # find ellipse
        # contours, _ = cv2.findContours(only_droplet_edge, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE) ## Does not work with troncated droplet
        contours = np.nonzero(only_droplet_edge==255)
        contours = (np.transpose((contours[1],contours[0])),)   ## done to switch the coordonates : [(x,y),..] to [(y,x),..] to have the same patern as cv2.findContours

        minEllipse = [None]*len(contours)
        for i, c in enumerate(contours):
            if c.shape[0] > 5:
                minEllipse[i] = cv2.fitEllipse(c)
        droplet_with_ellipse = copy.deepcopy(only_droplet_edge)
        for i, c in enumerate(contours):
            cv2.ellipse(droplet_with_ellipse, minEllipse[i], 127, 2)

        # save and display 
        if self.analysis_param["get_ellipse"]["save_with_edges"] or self.analysis_param["get_ellipse"]["display_with_edges"]:

            # image creation 
            droplet_with_ellipse = copy.deepcopy(only_droplet_edge)
            for i, c in enumerate(contours):
                cv2.ellipse(droplet_with_ellipse, minEllipse[i], 127, 2)

            if self.analysis_param["get_ellipse"]["save_with_edges"]:
                os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/ellipse", exist_ok=True)
                save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/ellipse/" + self.get_name_image(image_idx) +  "_ellipse.png"
                self.save_image(droplet_with_ellipse, save_name)
            if self.analysis_param["get_ellipse"]["display_with_edges"]:
                plt.figure()
                plt.imshow(droplet_with_ellipse) 

        return minEllipse

    def get_measurements_1image_1model(self, image_idx, model_idx):
        ''' image_idx : index of the image to analyse in the list self.images
            return the values of a (half of the length of the interface), h (height of the droplet) and R (radius) '''
        
        pred, image = self.predict_1image_1model(image_idx, model_idx, save_display=True, return_image=True)
        pred = pred.cpu().numpy()
        image = image.cpu().permute(1, 2, 0).numpy()

        pred_back = np.zeros_like(pred).astype('uint8')
        pred_droplet = np.zeros_like(pred).astype('uint8')
        pred_wire = np.zeros_like(pred).astype('uint8')

        pred_back[pred == 0] = 1
        pred_wire[pred == 1] = 1
        pred_droplet[pred == 2] = 1

        # Get the interface 
        droplet_edge = erode_and_dilate(pred_droplet)
        wire_edge = erode_and_dilate(pred_wire)
        interface = get_interface(wire_edge, droplet_edge)

        # Get the three points on the interface
        try: 
            a, b, x0, x1 = detect_lines(interface, display=False)
        except:
            print("unable to detect line for image with index ", image_idx)
            return [None, None, None]
        try:
            (x1_edge,y1_edge),(x2_edge,y2_edge),(x_center,y_center) = get_points(interface, x0, x1)
        except:
            print("unable to get point for image with index ", image_idx)
            return [None, None, None]

        # Get the perpendicular and points of the droplet
        try:
            a_perp, b_perp = get_perpendicular(a, b ,x_center, y_center, interface.shape[1])
            (x1_perp_drop_start, y1_perp_drop_start), (x2_perp_drop_end, y2_perp_drop_end), (x1_perp, y1_perp), (x2_perp, y2_perp) = get_droplet_points(a_perp, b_perp, droplet_edge)
            # print((x1_perp_drop_start, y1_perp_drop_start), (x2_perp_drop_end, y2_perp_drop_end))
        except:
            print("unable to detect perpendicular line for image with index ", image_idx)
            return [None, None, None]

        # Get the values
        try:
            R, x, y = self.get_circle(image_idx, model_idx, False)
        except:
            print("unable to get circle for image with index ", image_idx)
            return [None, None, None]
        a_cap = np.sqrt((y1_edge - y2_edge)**2 + (x1_edge - x2_edge)**2) / 2
        h = np.sqrt((y1_perp_drop_start - y2_perp_drop_end)**2 + (x1_perp_drop_start - x2_perp_drop_end)**2)
        # print(h)
        
        if h < 30:
            # print("h computed with the radius")
            diff_h_R = np.sqrt((y1_perp_drop_start - y)**2 + (x1_perp_drop_start - x)**2)
            h = diff_h_R + R

        if self.analysis_param["measurements"]["save_with_lines"] or self.analysis_param["measurements"]["display_with_lines"]:
            plt.figure()
            plt.xlim(0, image.shape[0])
            plt.ylim((image.shape[1], 0))
            plt.imshow(image, cmap='Greys_r')
            plt.plot((x0, x1), (0, interface.shape[1]), '-b',  linewidth=1)
            plt.scatter((x1_edge, x2_edge, x_center), (y1_edge, y2_edge, y_center), s=50)
            plt.scatter(x,y, s=50)
            plt.scatter((x1_perp_drop_start, x2_perp_drop_end), (y1_perp_drop_start, y2_perp_drop_end), s=50)
            plt.plot((x1_perp_drop_start, x2_perp_drop_end), (y1_perp_drop_start, y2_perp_drop_end), '-r', linewidth=1)
            if self.analysis_param["measurements"]["save_with_lines"]:
                os.makedirs(self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/lines", exist_ok=True)
                save_name = self.main_analysis_dir + "/prediction_" + str(self.models_names[model_idx]) + "/lines/" + self.get_name_image(image_idx) + "_lines.png" 
                plt.savefig(save_name)
                plt.close()
            if self.analysis_param["measurements"]["display_with_lines"]:
                plt.show()

        return [a_cap, h, R]
    
    def get_measurements_1image(self, image_idx, reduction='mean'):
        ''' return a list of 3-uplet-list [a, h, R]. See method get_measurements_1image_1model for more details'''
        res = [ self.get_measurements_1image_1model(image_idx, model_idx) for model_idx in range(len(self.models)) ]
        res = [ x for x in res if x != [None, None, None] ]
        if reduction == 'mean':
            if res != []:
                res = np.array(res)
                return np.mean(res, axis=0)
            else:
                return [None, None, None]
        if reduction == 'none':
            return res
        
    def create_result_file(self, result_path):
        # if self.result_file is None:
        workbook = openpyxl.Workbook()
        worksheet = workbook[workbook.active.title]
        worksheet["A1"] = "frame"
        worksheet["B1"] = "a (half length of the interface)"
        worksheet["C1"] = "h (height of the droplet)"
        worksheet["D1"] = "R (radius)"
        # name = "D:/Images_nanomax/Contact_angle/result_demo.xlsx" ## Change here !!! 
        workbook.save(result_path)
        self.result_file = result_path
        # else:
        #     raise RuntimeError("Issue - the excel result file already exists - path:", self.result_file)

    def measure_and_write_all(self, result_path):
        ''' Normally, it is the first time values are writen in the result file '''
        self.create_result_file(result_path)
        workbook = openpyxl.load_workbook(self.result_file)
        worksheet = workbook[workbook.active.title]

        for idx in range(len(self.images)):
            values = self.get_measurements_1image(idx, reduction='mean')
            # writing frame
            worksheet.cell(row=idx+2, column=1).value = idx + 1
            # writing a
            worksheet.cell(row=idx+2, column=2).value = values[0]
            # writing h
            worksheet.cell(row=idx+2, column=3).value = values[1]
            # writing R
            worksheet.cell(row=idx+2, column=4).value = values[2]
            if idx % 100 == 0:
                print(idx)
        workbook.save(self.result_file)

    def compute_write_contact_angle(self):
        ''' assuming that the values of a, h and R are already writen in the result file '''
        assert self.result_file is not None
        workbook = openpyxl.load_workbook(self.result_file)
        worksheet = workbook[workbook.active.title]
        worksheet["E1"] = "contact angle v1 (rad)"
        # worksheet["F1"] = "contact angle v2 (rad)"
        # worksheet["G1"] = "contact angle v3 (rad)"

        for idx in range(len(self.images)):
            a = worksheet.cell(row=idx+2, column=2).value
            h = worksheet.cell(row=idx+2, column=3).value
            R = worksheet.cell(row=idx+2, column=4).value
            # writing v1
            worksheet.cell(row=idx+2, column=5).value = get_contact_angle_v1(h, a)
            # writing v2
            # worksheet.cell(row=idx+2, column=6).value = get_contact_angle_v2(R, h)
            # writing v3
            # worksheet.cell(row=idx+2, column=7).value = get_contact_angle_v3(a, R)
            if idx % 100 == 0:
                print(idx)
        workbook.save(self.result_file)

    def do_the_measurement(self, result_file):
        self.measure_and_write_all(result_file)
        print("measurement done")
        self.compute_write_contact_angle()
        print("done")

    def get_variable_name(self, variable, own=True):
        ''' df_comparaison is based on the template of the file tablev1'. 
            df_own is based on the template defined by the methods of this class (see create_result_file). '''
        df_own = {"contact angle v1" : "contact angle v1 (rad)", 
                    "contact angle v2" : "contact angle v2 (rad)", 
                    "contact angle v3" : "contact angle v3 (rad)",
                    "interface" : "a (half length of the interface)", 
                    "height" : "h (height of the droplet)",
                    "radius" : "R (radius)"}
        df_comparaison = {"contact angle v1" : "contact angle (°)", 
                            "contact angle v2" : "contact angle (°)",
                            "contact angle v3" : "contact angle (°)",
                            "interface" : "diameter (nm)", 
                            "height" : "height (nm)"}
        try: 
            if own:
                return df_own[variable]
            else:
                return df_comparaison[variable]
        except KeyError:
            print("variable " + variable + " not available")
            raise KeyError("variable not available")

    def get_values_from_xls_file(self, xls_path, variable, own, convert_degree=True):
        ''' extract the values of variable from an excel file.
            The column and the rows must follow the right template: 
            file 'tablev1' for own=False and file created by the methods of this class for own=True.
            Returns two lists, the first one is frames, the second one is values '''
        if own:
            df = pd.read_excel(self.result_file)
        else:   
            df = pd.read_excel(xls_path)

        variable_list = []
        frame_list = []
        try:
            variable = self.get_variable_name(variable, own)
        except KeyError:
            print("variable name not found")
            return [], []

        for index, row in df.iterrows():
            if np.isnan(row['frame']):
                continue
            frame = int(row['frame'])
            if not(np.isnan(row[variable])):
                variable_list.append(row[variable])
                frame_list.append(frame)
        if convert_degree:
            variable_list = 180 * np.array(variable_list) / np.pi
        return frame_list, variable_list
    
    def plot_variable(self, variable, comparaison_file):
        # get the values :
        frame_own, variable_own = self.get_values_from_xls_file(self.result_file, variable, own=True, convert_degree=True)
        if comparaison_file:
            frame_comparaison, variable_comparaison = self.get_values_from_xls_file(comparaison_file, variable, own=False, convert_degree=False)
        
        plt.figure(figsize=(30, 15))
        if comparaison_file:
            plt.plot(frame_comparaison, variable_comparaison, 'o-', label="measured values", color="r")
        plt.scatter(frame_own, variable_own, label="predicted values")
        plt.grid()
        plt.title("Evolution of " + variable)
        plt.legend()
        if variable.startswith("contact angle"):
            plt.ylim((98,120))
        plt.show()

    def mse_xls_file(self, comparaison_file):
        wbA = openpyxl.load_workbook(comparaison_file)
        wbB = openpyxl.load_workbook(self.result_file)
        sheetA = wbA.active
        sheetB = wbB.active

        frame_col_idx_A = get_column_index(sheetA, "frame")
        angle_col_idx_A = get_column_index(sheetA, "contact angle (°)")

        frame_col_idx_B = get_column_index(sheetB, "frame")
        angle_col_idx_B = get_column_index(sheetB, "contact angle v1 (rad)")

        dataA = {}
        for row in sheetA.iter_rows(min_row=2, values_only=True):
            frame_index, contact_angle = row[frame_col_idx_A - 1], row[angle_col_idx_A - 1]
            dataA[frame_index] = contact_angle

        dataB = {}
        for row in sheetB.iter_rows(min_row=2, values_only=True):
            frame_index, contact_angle = row[frame_col_idx_B - 1], row[angle_col_idx_B - 1]
            dataB[frame_index] = contact_angle

        mse = 0.0
        count = 0

        for frame_index, angleA in dataA.items():
            if angleA is not None:
                angleB = dataB.get(frame_index)
                if angleB is not None:
                    angleB = 180 * angleB / np.pi
                    mse += (angleA - angleB) ** 2
                    count += 1

        # Calculate the mean square error
        if count > 0:
            mse /= count
        print(count)

        print(f"Mean Square Error (MSE): {mse}")

        # Close the Excel files
        wbA.close()
        wbB.close()

        return mse


#%% tests 

model_folder = "C:/Users/lucas/OneDrive/C2N/trained_models"
models = [ "model_super_20230723_155619", 
          "model_semi_20230719_134837", 
          "model_semi_20230725_094455", 
          "unet_super_20230807_081324", 
          "unet_semi_20230807_094900", 
          "unet_super_20230812_230324"
          ]

# images = "C:/Users/lucas/OneDrive/C2N/Contact angle/full"
# image_folder = "C:/Users/lucas/OneDrive/C2N/Images/unlabeled_data/unlabeled_images_512_pt_10000"
# files = os.listdir(image_folder)
# file = random.choice(files)
# print(file)
# image_test =  image_folder + "/" + file

for model in models:
    print(model)
    ### for the paper analysis 
    # test = Analysis([model], images_path=images, videos_path=None,  result_file=None)
    # test.predict_1image_1model(image_idx=0, model_idx=0)
    # res_file_name = "C:/Users/lucas/OneDrive/C2N/results_paper/res_model_" + model + ".xlsx" 
    # test.do_the_measurement(res_file_name)

    ### just for the prediction 
    # test = Analysis([model], images_path=image_test, videos_path=None,  result_file=None)
    # test.predict_1image_1model(image_idx=0, model_idx=0)

    ## plot comparaison
    images = ""
    own_path = "C:/Users/lucas/OneDrive/C2N/results_paper/res_model_" + model + ".xlsx"
    comp_path = "C:/Users/lucas/OneDrive/C2N/results_paper/results_nam.xlsx"
    test = Analysis([model], images, None, result_file=own_path)
    test.mse_xls_file(comp_path)














# test.get_ellipse(image_idx=194, model_idx=0)
# test.get_measurements_1image_1model(image_idx=194, model_idx=0)

# name = "D:/Images_segmentation/ellipse2/result_demo.xlsx"
# test.do_the_measurement(name)

# model_folder = "C:/Users/lucas.degeorge/Documents/trained_models"
# models = [model_folder + "/unet_20230828_094619/unet_20230828_094619_best.pth"]

# images = "D:/Images_segmentation/Ellipse/frames/6_frames/6_frame_00142.png"
# images = "D:/Images_segmentation/Ellipse/pseudo_training/to_segment/6_frame_175.png"
# test = Analysis(models, images, None, 512)
# test.get_circle(0, 0)
# for i in range(len(os.listdir(images))):
# A = test.get_ellipse(0, 0)
# test.do_the_measurement()





# images_path = "D:/Images_segmentation/Ellipse/frames"
# images_list = [ images_path + "/5_frames", images_path + "/6_frames", images_path + "/13_frames", images_path + "/14_frames" ]
# for path in images_list:
#     print(path)
#     test = Analysis(models, path, None, 512)
#     test.predict(0)

# test = Analysis(models, images_path + "/6_frames", None, 512)
# test.predict(0)

 
# images = ""
# own_path = "D:/Images_segmentation/Contact_angle/result_test.xlsx"
# comp_path = "D:/Images_segmentation/Contact_angle/tablev1.xls"
# test = Analysis(models, images, None, 512, result_file=own_path)
# test.plot_variable("contact angle v1", comp_path)

